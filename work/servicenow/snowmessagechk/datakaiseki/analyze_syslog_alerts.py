#!/usr/bin/env python3
"""
em_alert テーブルから syslog ソースのアラートを取得し、
description の形式ごとに分類、CI バインド状況を付与して Excel 出力する。

対象: source=syslog かつ sys_created_on >= 2026-04-01

出力:
  datakaiseki/syslog_alerts_YYYYMMDD.xlsx
    - 概要         : 件数・CI バインド率・パターン別サマリー
    - 全件一覧      : 全レコード（パターンラベル付き）
    - パターン別シート: パターンごとのレコード一覧
"""

import sys
import os
import re
import json
import datetime
from collections import Counter, defaultdict

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ["SNOW_BASE_URL"]    = "https://biglobeprod.service-now.com"
os.environ["SNOW_SECRET_NAME"] = "servicenow/api/credentials/biglobeprod/admin-ai-api"
os.environ["SNOW_CLIENT_ID"]   = ""
os.environ["SNOW_CLIENT_SECRET"] = ""

import snow_client

FETCH_FIELDS = [
    "sys_id", "number", "source", "node", "type", "resource",
    "severity", "description", "message_key", "additional_info",
    "cmdb_ci", "classification", "u_type_category",
    "u_matched_rules", "sys_created_on",
]

PAGE_SIZE = 1000

# description の形式パターン（上から順に評価、最初にマッチしたものを採用）
PATTERNS: list[tuple[str, re.Pattern]] = [
    ("OSPF",              re.compile(r"\bOSPF\b",              re.I)),
    ("BGP",               re.compile(r"\bBGP\b",               re.I)),
    ("ISIS/IS-IS",        re.compile(r"\bIS-?IS\b",            re.I)),
    ("LDP",               re.compile(r"\bLDP\b",               re.I)),
    ("RSVP",              re.compile(r"\bRSVP\b",              re.I)),
    ("MPLS",              re.compile(r"\bMPLS\b",              re.I)),
    ("Interface Up/Down", re.compile(r"\binterface\b.{0,50}(up|down)\b", re.I)),
    ("Link Up/Down",      re.compile(r"\blink\b.{0,50}(up|down)\b",      re.I)),
    ("Interface Error",   re.compile(r"\b(interface|port)\b.{0,50}err", re.I)),
    ("Chassis/Hardware",  re.compile(r"\b(chassis|hardware|FPC|PIC|MIC|RE)\b", re.I)),
    ("FAN/Temperature",   re.compile(r"\b(fan|temp|thermal|overheat)\b", re.I)),
    ("Power",             re.compile(r"\b(power|PSU|voltage)\b",          re.I)),
    ("Memory/CPU",        re.compile(r"\b(memory|cpu|load|util)\b",       re.I)),
    ("Alarm",             re.compile(r"\balarm\b",                        re.I)),
    ("ACL/Firewall",      re.compile(r"\b(acl|firewall|filter|deny)\b",  re.I)),
    ("Login/Auth",        re.compile(r"\b(login|logout|auth|ssh|telnet)\b", re.I)),
    ("Config Change",     re.compile(r"\b(config|commit|change|edit)\b",  re.I)),
    ("SNMP",              re.compile(r"\bSNMP\b",                         re.I)),
    ("Syslog Marker",     re.compile(r"MARK|newsyslog",                   re.I)),
    ("(空)",              re.compile(r"^$")),
]


def classify(desc: str) -> str:
    d = (desc or "").strip()
    if not d:
        return "(空)"
    for label, pat in PATTERNS:
        if pat.search(d):
            return label
    return "その他"


def _gs(rec: dict, fname: str) -> str:
    val = rec.get(fname, "")
    if isinstance(val, dict):
        return str(val.get("display_value") or val.get("value") or "")
    return str(val or "")


_RE_IP        = re.compile(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$")
_RE_GW_FPC    = re.compile(r"^(.*gw.*)-fpc\d+$",  re.I)   # Pattern A
_RE_GW_RE     = re.compile(r"^(.*gw.*)-re\d+$",   re.I)   # Pattern C
_RE_FPC_ALONE = re.compile(r"^fpc\d+$",            re.I)   # Pattern B


def _is_ip(node: str) -> bool:
    return bool(_RE_IP.match(node or ""))


def detect_remedy_pattern(node: str) -> tuple[str, str]:
    """(パターン記号, 正規化後node名) を返す。
    A: gw系FPCサフィックス混入  →  -fpcN 除去
    B: fpcN 単体（無視）
    C: gw系REサフィックス混入   →  -reN 除去
    D: IPアドレス形式
    """
    if not node:
        return ("", "")
    if _RE_FPC_ALONE.match(node):
        return ("B", "")
    m = _RE_GW_FPC.match(node)
    if m:
        return ("A", m.group(1))
    m = _RE_GW_RE.match(node)
    if m:
        return ("C", m.group(1))
    if _is_ip(node):
        return ("D", node)
    return ("", "")


def lookup_cmdb_by_ip(token: str, ips: set[str]) -> tuple[dict[str, dict], dict[str, dict]]:
    """IP アドレスで CMDB CI を照合する。
    戻り値: (private_map, public_map)  各 {ip: cmdb_ci_record}
    """
    ip_list = sorted(ips)
    chunk_size = 80
    private_map: dict[str, dict] = {}
    public_map:  dict[str, dict] = {}

    fields = "sys_id,name,sys_class_name,u_private_ip_address,u_public_ip_address"

    print(f"  IP CMDB照合: {len(ip_list)} 件", file=sys.stderr)
    for i in range(0, len(ip_list), chunk_size):
        chunk = ip_list[i:i + chunk_size]
        escaped = [ip.replace(",", r"\,") for ip in chunk]
        ip_str = ",".join(escaped)
        for field_name, target_map in [
            ("u_private_ip_address", private_map),
            ("u_public_ip_address",  public_map),
        ]:
            try:
                recs = snow_client.table_get(token, "cmdb_ci", {
                    "sysparm_limit":         500,
                    "sysparm_query":         f"{field_name}IN{ip_str}",
                    "sysparm_display_value": "true",
                    "sysparm_fields":        fields,
                })
                for rec in recs:
                    ip_val = _gs(rec, field_name)
                    if ip_val:
                        target_map[ip_val] = rec
            except Exception as exc:
                print(f"  CMDB IP照合エラー ({field_name}): {exc}", file=sys.stderr)

    found = len(set(private_map) | set(public_map))
    print(f"  IP CMDB照合完了: private={len(private_map)}件 / public={len(public_map)}件 / 合計ユニーク={found}件", file=sys.stderr)
    return private_map, public_map


def lookup_cmdb_by_names(token: str, names: set[str]) -> dict[str, dict]:
    """cmdb_ci_ip_router を CI名で一括照合する。戻り値: {name: record}"""
    result: dict[str, dict] = {}
    name_list = sorted(names)
    chunk_size = 80
    print(f"  名前 CMDB照合(ip_router): {len(name_list)} 件", file=sys.stderr)
    for i in range(0, len(name_list), chunk_size):
        chunk = name_list[i:i + chunk_size]
        escaped = [n.replace(",", r"\,") for n in chunk]
        try:
            recs = snow_client.table_get(token, "cmdb_ci_ip_router", {
                "sysparm_limit":         500,
                "sysparm_query":         "nameIN" + ",".join(escaped),
                "sysparm_display_value": "true",
                "sysparm_fields":        "sys_id,name,sys_class_name",
            })
            for rec in recs:
                name = _gs(rec, "name")
                if name:
                    result[name] = rec
        except Exception as exc:
            print(f"  CMDB名前照合エラー: {exc}", file=sys.stderr)
    print(f"  名前照合完了: {len(result)} 件ヒット", file=sys.stderr)
    return result


def fetch_syslog_alerts(token: str) -> list[dict]:
    all_recs: list[dict] = []
    offset = 0
    query = "source=syslog^sys_created_on>=2026-04-01 00:00:00"
    print(f"  クエリ: {query}", file=sys.stderr)
    while True:
        recs = snow_client.table_get(token, "em_alert", {
            "sysparm_limit":         PAGE_SIZE,
            "sysparm_offset":        offset,
            "sysparm_query":         query,
            "sysparm_display_value": "true",
            "sysparm_fields":        ",".join(FETCH_FIELDS),
        })
        if not recs:
            break
        all_recs.extend(recs)
        print(f"  取得済み: {len(all_recs)} 件", file=sys.stderr)
        if len(recs) < PAGE_SIZE:
            break
        offset += PAGE_SIZE
    return all_recs


def output_excel(recs: list[dict],
                 ip_private_map: dict[str, dict],
                 ip_public_map:  dict[str, dict],
                 name_ci_map:    dict[str, dict],
                 filepath: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl が必要です: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb = openpyxl.Workbook()

    title_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    hdr_fill   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ok_fill    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ng_fill    = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, size=13)
    hdr_font   = Font(color="FFFFFF", bold=True)
    wrap       = Alignment(wrap_text=True, vertical="top")

    def _hdr(ws, row: int, col: int, val: str) -> None:
        c = ws.cell(row=row, column=col, value=val)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    # パターン分類・CI バインド状況・IP CI照合を付与
    for rec in recs:
        rec["_pattern"]  = classify(_gs(rec, "description"))
        ci_val = rec.get("cmdb_ci", "")
        if isinstance(ci_val, dict):
            ci_name = ci_val.get("display_value") or ""
            ci_id   = ci_val.get("value") or ""
        else:
            ci_name = str(ci_val or "")
            ci_id   = ""
        rec["_ci_name"]  = ci_name
        rec["_ci_id"]    = ci_id
        rec["_ci_bound"] = "あり" if ci_name else "なし"

        # IP アドレス node の CMDB 照合
        node = _gs(rec, "node")
        if _is_ip(node):
            rec["_node_is_ip"] = "○"
            priv_ci = ip_private_map.get(node)
            pub_ci  = ip_public_map.get(node)
            rec["_ip_ci_private"] = _gs(priv_ci, "name") if priv_ci else ""
            rec["_ip_ci_public"]  = _gs(pub_ci,  "name") if pub_ci  else ""
            rec["_ip_ci_class"]   = (
                _gs(priv_ci, "sys_class_name") or _gs(pub_ci, "sys_class_name")
                if (priv_ci or pub_ci) else ""
            )
            rec["_ip_ci_found"]   = "あり" if (priv_ci or pub_ci) else "なし"
        else:
            rec["_node_is_ip"]    = ""
            rec["_ip_ci_private"] = ""
            rec["_ip_ci_public"]  = ""
            rec["_ip_ci_class"]   = ""
            rec["_ip_ci_found"]   = ""

        # 対処パターン判定
        pat, norm = detect_remedy_pattern(node)
        rec["_remedy_pattern"] = pat
        rec["_remedy_norm_node"] = norm
        if pat == "B":
            rec["_remedy_ci"]    = ""
            rec["_remedy_class"] = ""
            rec["_remedy_found"] = "無視"
            rec["_remedy_desc"]  = "fpcN単体ノード → 無視"
        elif pat == "A":
            ci = name_ci_map.get(norm)
            rec["_remedy_ci"]    = _gs(ci, "name")    if ci else ""
            rec["_remedy_class"] = _gs(ci, "sys_class_name") if ci else ""
            rec["_remedy_found"] = "あり" if ci else "なし"
            rec["_remedy_desc"]  = (
                f"FPCサフィックス除去({norm}) → CI特定可: {_gs(ci,'name')} [{_gs(ci,'sys_class_name')}]"
                if ci else
                f"FPCサフィックス除去({norm}) → CI未特定（ip_routerに登録なし）"
            )
        elif pat == "C":
            ci = name_ci_map.get(norm)
            rec["_remedy_ci"]    = _gs(ci, "name")    if ci else ""
            rec["_remedy_class"] = _gs(ci, "sys_class_name") if ci else ""
            rec["_remedy_found"] = "あり" if ci else "なし"
            rec["_remedy_desc"]  = (
                f"REサフィックス除去({norm}) → CI特定可: {_gs(ci,'name')} [{_gs(ci,'sys_class_name')}]"
                if ci else
                f"REサフィックス除去({norm}) → CI未特定（ip_routerに登録なし）"
            )
        elif pat == "D":
            priv_ci = ip_private_map.get(node)
            pub_ci  = ip_public_map.get(node)
            ci = priv_ci or pub_ci
            rec["_remedy_ci"]    = _gs(ci, "name")    if ci else ""
            rec["_remedy_class"] = _gs(ci, "sys_class_name") if ci else ""
            rec["_remedy_found"] = "あり" if ci else "なし"
            if ci:
                mf = "u_private_ip_address" if priv_ci else "u_public_ip_address"
                rec["_remedy_desc"] = (
                    f"IP照合({mf}) → CI特定可: {_gs(ci,'name')} [{_gs(ci,'sys_class_name')}]"
                )
            else:
                rec["_remedy_desc"] = "IP照合 → CI未特定（ip_router/インタフェースに登録なし）"
        else:
            rec["_remedy_ci"]    = ""
            rec["_remedy_class"] = ""
            rec["_remedy_found"] = ""
            rec["_remedy_desc"]  = ""

    by_pattern: dict[str, list[dict]] = defaultdict(list)
    for rec in recs:
        by_pattern[rec["_pattern"]].append(rec)

    bound_count   = sum(1 for r in recs if r["_ci_bound"] == "あり")
    unbound_count = len(recs) - bound_count

    # ================================================================
    # 概要シート
    # ================================================================
    ws0 = wb.active
    ws0.title = "概要"

    cell = ws0.cell(row=1, column=1,
                    value="syslog アラート 分析レポート（4月以降）")
    cell.fill = title_fill
    cell.font = title_font

    ws0.cell(row=2, column=1, value="取得日時")
    ws0.cell(row=2, column=2, value=now_str)
    ws0.cell(row=3, column=1, value="対象期間")
    ws0.cell(row=3, column=2, value="2026-04-01 以降")
    ws0.cell(row=4, column=1, value="総件数")
    ws0.cell(row=4, column=2, value=len(recs))
    ws0.cell(row=5, column=1, value="CI バインド済み")
    ws0.cell(row=5, column=2, value=bound_count)
    ws0.cell(row=6, column=1, value="CI 未バインド")
    ws0.cell(row=6, column=2, value=unbound_count)

    ws0.cell(row=8, column=1, value="パターン別サマリー").font = Font(bold=True)
    _hdr(ws0, 9, 1, "パターン")
    _hdr(ws0, 9, 2, "件数")
    _hdr(ws0, 9, 3, "CI バインド済み")
    _hdr(ws0, 9, 4, "CI 未バインド")
    _hdr(ws0, 9, 5, "バインド率")
    for r_idx, (pat, precs) in enumerate(
            sorted(by_pattern.items(), key=lambda x: -len(x[1])), 10):
        bc = sum(1 for r in precs if r["_ci_bound"] == "あり")
        uc = len(precs) - bc
        rate = f"{bc / len(precs) * 100:.1f}%" if precs else ""
        ws0.cell(row=r_idx, column=1, value=pat)
        ws0.cell(row=r_idx, column=2, value=len(precs))
        ws0.cell(row=r_idx, column=3, value=bc)
        ws0.cell(row=r_idx, column=4, value=uc)
        ws0.cell(row=r_idx, column=5, value=rate)

    ws0.column_dimensions["A"].width = 25
    ws0.column_dimensions["B"].width = 10
    ws0.column_dimensions["C"].width = 14
    ws0.column_dimensions["D"].width = 12
    ws0.column_dimensions["E"].width = 12

    # ================================================================
    # 全件一覧シート
    # ================================================================
    LIST_COLS = [
        ("パターン",              "_pattern"),
        ("CI_バインド",           "_ci_bound"),
        ("CI名",                 "_ci_name"),
        # --- IP照合列 ---
        ("node_is_IP",           "_node_is_ip"),
        ("IP_CI特定",            "_ip_ci_found"),
        ("IP_CI候補(private_ip)","_ip_ci_private"),
        ("IP_CI候補(public_ip)", "_ip_ci_public"),
        ("IP_CIクラス",          "_ip_ci_class"),
        # --- 対処法列 ---
        ("対処パターン",          "_remedy_pattern"),
        ("対処_CI特定",          "_remedy_found"),
        ("対処後_CI候補",         "_remedy_ci"),
        ("対処後_CIクラス",       "_remedy_class"),
        ("対処法",               "_remedy_desc"),
        # --- アラートデータ ---
        ("number",               "number"),
        ("node",                 "node"),
        ("正規化node",            "_remedy_norm_node"),
        ("severity",             "severity"),
        ("description",          "description"),
        ("message_key",          "message_key"),
        ("u_matched_rules",      "u_matched_rules"),
        ("sys_created_on",       "sys_created_on"),
        ("sys_id",               "sys_id"),
    ]

    ws_all = wb.create_sheet("全件一覧")
    for col, (label, _) in enumerate(LIST_COLS, 1):
        _hdr(ws_all, 1, col, label)

    from openpyxl.utils import get_column_letter
    _COLOR_LABELS = {"CI_バインド", "IP_CI特定", "対処_CI特定"}
    COL_WIDTHS = [
        20, 12, 28,
        10, 12, 28, 28, 22,
        12, 12, 28, 22, 55,
        14, 30, 20,
        10, 60, 50, 40, 18, 36,
    ]

    for r_idx, rec in enumerate(
            sorted(recs, key=lambda x: x["_pattern"]), 2):
        for col, (label, fname) in enumerate(LIST_COLS, 1):
            val = rec.get(fname, "") if fname.startswith("_") else _gs(rec, fname)
            cell = ws_all.cell(row=r_idx, column=col, value=str(val or "")[:500])
            if label in _COLOR_LABELS and val in ("あり", "なし"):
                cell.fill = ok_fill if val == "あり" else ng_fill

    for i, w in enumerate(COL_WIDTHS, 1):
        ws_all.column_dimensions[get_column_letter(i)].width = w

    # ================================================================
    # パターン別シート
    # ================================================================
    def _safe_sheet_name(name: str) -> str:
        invalid = r'\/:*?"<>|'
        s = "".join(c for c in name if c not in invalid) or "unnamed"
        return s[:31]

    for pat, precs in sorted(by_pattern.items(), key=lambda x: -len(x[1])):
        sname = _safe_sheet_name(pat)
        existing = {s.title for s in wb.worksheets}
        if sname in existing:
            sname = sname[:28] + "_2"
        ws = wb.create_sheet(sname)

        # タイトル
        cell = ws.cell(row=1, column=1,
                       value=f"【{pat}】  {len(precs)} 件")
        cell.fill = title_fill
        cell.font = title_font

        # サマリー
        bc = sum(1 for r in precs if r["_ci_bound"] == "あり")
        ws.cell(row=3, column=1, value="件数")
        ws.cell(row=3, column=2, value=len(precs))
        ws.cell(row=4, column=1, value="CI バインド済み")
        ws.cell(row=4, column=2, value=bc)
        ws.cell(row=5, column=1, value="CI 未バインド")
        ws.cell(row=5, column=2, value=len(precs) - bc)

        # node 別件数
        node_cnt = Counter(_gs(r, "node") or "(空)" for r in precs)
        ws.cell(row=7, column=1, value=f"node別（{len(node_cnt)}種）").font = Font(bold=True)
        _hdr(ws, 8, 1, "node")
        _hdr(ws, 8, 2, "件数")
        _hdr(ws, 8, 3, "CI バインド済み件数")
        for ri, (node, cnt) in enumerate(node_cnt.most_common(50), 9):
            bc_node = sum(1 for r in precs
                          if (_gs(r, "node") or "(空)") == node
                          and r["_ci_bound"] == "あり")
            ws.cell(row=ri, column=1, value=node)
            ws.cell(row=ri, column=2, value=cnt)
            ws.cell(row=ri, column=3, value=bc_node)

        # description サンプル（最大5種）
        sample_row = 9 + min(len(node_cnt), 50) + 2
        ws.cell(row=sample_row, column=1,
                value="description サンプル（最大5件）").font = Font(bold=True)
        seen_desc: set[str] = set()
        sr = sample_row + 1
        for rec in precs:
            desc = _gs(rec, "description")
            if desc not in seen_desc:
                seen_desc.add(desc)
                ws.cell(row=sr, column=1, value=desc[:300])
                ws.cell(row=sr, column=1).alignment = wrap
                sr += 1
            if len(seen_desc) >= 5:
                break

        # 詳細一覧
        detail_row = sr + 2
        ws.cell(row=detail_row, column=1, value="詳細一覧").font = Font(bold=True)
        for col, (label, _) in enumerate(LIST_COLS, 1):
            _hdr(ws, detail_row + 1, col, label)
        for ri2, rec in enumerate(precs, detail_row + 2):
            for col, (label, fname) in enumerate(LIST_COLS, 1):
                val = rec.get(fname, "") if fname.startswith("_") else _gs(rec, fname)
                cell = ws.cell(row=ri2, column=col, value=str(val or "")[:500])
                if label in _COLOR_LABELS and val in ("あり", "なし"):
                    cell.fill = ok_fill if val == "あり" else ng_fill

        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(filepath)
    print(f"Excel 出力完了: {filepath}", file=sys.stderr)


def main() -> None:
    now = datetime.datetime.now()
    filepath = now.strftime(
        os.path.join(os.path.dirname(__file__), "syslog_alerts_%Y%m%d.xlsx")
    )

    print("OAuth トークンを取得中...", file=sys.stderr)
    token = snow_client.get_token()

    print("syslog アラートを取得中（4月以降）...", file=sys.stderr)
    recs = fetch_syslog_alerts(token)
    print(f"合計取得: {len(recs)} 件", file=sys.stderr)

    if not recs:
        print("対象レコードが0件でした。", file=sys.stderr)
        sys.exit(0)

    # IP アドレス形式の node を抽出して CMDB 照合（Pattern D）
    ip_nodes = {_gs(r, "node") for r in recs if _is_ip(_gs(r, "node"))}
    print(f"  IPアドレス形式 node: {len(ip_nodes)} 種", file=sys.stderr)
    if ip_nodes:
        ip_private_map, ip_public_map = lookup_cmdb_by_ip(token, ip_nodes)
    else:
        ip_private_map, ip_public_map = {}, {}

    # Pattern A/C: FPC/RE サフィックス除去後の正規化node名で cmdb_ci_ip_router を照合
    norm_names: set[str] = set()
    for r in recs:
        pat, norm = detect_remedy_pattern(_gs(r, "node"))
        if pat in ("A", "C") and norm:
            norm_names.add(norm)
    print(f"  正規化node（パターンA/C）: {len(norm_names)} 種", file=sys.stderr)
    name_ci_map = lookup_cmdb_by_names(token, norm_names) if norm_names else {}

    output_excel(recs, ip_private_map, ip_public_map, name_ci_map, filepath)


if __name__ == "__main__":
    main()
