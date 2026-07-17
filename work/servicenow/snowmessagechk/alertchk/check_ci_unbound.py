#!/usr/bin/env python3
"""
ServiceNow em_alert テーブルの CI未バインドアラートを調査する。

ci_unbound_alerts20260616.xlsx と同じ形式で出力する。
Zabbix / CloudWatchLogs / HIOS(SV) の各シート末尾に疑義CI CMDB照合表を追加する。

除外ソース（調査対象外）:
  既存除外: HIOS(AWS)、iMark_AWS(Servicekanshi)、キャリア障害系、Zabbixハートビート、
           業連メール、DDoS、JPIX、ウェザーニューズ、Downdetector
  追加除外: WebAI、Service Health Dashboard Alarm、Email、DeepField/Arbor vSP、
           ServiceNowテストメールアラート、EMSelfMonitoring、bousai、
           ServiceNow UATテストメールアラート

Excel 出力シート構成:
  1. 概要           - ソース別集計・除外条件
  2. CI未バインド一覧 - 除外後全件
  3. <source>       - ソース別分析シート（分布表 + CMDB照合表 ※3ソースのみ）
  N. 除外ソース      - 除外理由別件数

使用例:
  python check_ci_unbound.py
  python check_ci_unbound.py --file tmpdir/ci_unbound_alerts.xlsx
  python check_ci_unbound.py --no-cmdb   # CMDB照合スキップ（高速確認用）
  python check_ci_unbound.py --limit 5000
"""

import sys
import os
import json
import argparse
from collections import defaultdict, Counter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

os.environ["SNOW_BASE_URL"] = "https://biglobeprod.service-now.com"
os.environ["SNOW_SECRET_NAME"] = "servicenow/api/credentials/biglobeprod/admin-ai-api"
os.environ["SNOW_CLIENT_ID"] = ""
os.environ["SNOW_CLIENT_SECRET"] = ""

import snow_client

FETCH_FIELDS = [
    "sys_id", "source", "node", "type", "resource",
    "severity", "classification", "u_type_category",
    "u_matched_rules", "additional_info", "sys_created_on",
]

PAGE_SIZE = 1000

EXCLUDE_EXACT_SOURCES = frozenset({
    "業連メール",
    "Downdetector",
    "DDoS",
    "JPIX",
    "ウェザーニューズ",
    # 問題なし確認済み（調査不要）
    "WebAI",
    "Service Health Dashboard Alarm",
    "Email",
    "DeepField/Arbor vSP",
    "ServiceNowテストメールアラート",
    "EMSelfMonitoring",
    "bousai",
    "ServiceNow UATテストメールアラート",
    "ExpressList表示対象外",
    "工事連絡",
    # 対応不要
    "Mackerel",
})

# CMDB照合・シート末尾に照合表を追加するソース
CMDB_LOOKUP_SOURCES = frozenset({"Zabbix", "CloudWatchLogs", "HIOS(SV)", "HIOS(AWS)"})

# ソース別アノテーション（既知情報）
SOURCE_ANNOTATIONS: dict[str, dict] = {
    "Zabbix": {
        "priority": "高",
        "cause": "resourceが全件空。nodeにIF名(_et-0/1/2等)が混入しCMDB名と不一致。ssap/htap系はCMDB未登録。iMarkN系は外部URL監視でCI登録不可。",
        "action": "①gw/sw/bb/bgar系: matchRuleでnode名正規化（IF名除去）。②ssap/htap系: CMDB登録。③iMarkN系: CIなし処理対象として除外検討。",
    },
    "CloudWatchLogs": {
        "priority": "高",
        "cause": "typeフィールドが全件空。CloudWatchLogsソース向けのCIバインドルール（matchRule）が未設定。",
        "action": "em_rule_xml に CloudWatchLogs ソース向けCIバインドルール追加。em_mapping_rule でtype/resourceをフィールドマッピングする必要あり。",
    },
    "HIOS(AWS)": {
        "priority": "高",
        "cause": "CIなしで処理されていたが除外解除。nodeはAWS CloudWatchメトリクス系。CIバインドルールが未設定の可能性あり。",
        "action": "CIバインドルール設定状況を確認。nodeのCMDB登録状況を確認。",
    },
    "HIOS(SV)": {
        "priority": "高",
        "cause": "type/resource全件空。nodeはbgeb/bvaw系（CMDB登録済み）とw19ad系（未登録）。HIOS(SV)ソース向けCIバインドルールが未設定。severity全件重要。",
        "action": "HIOS(SV)ソース向けCIバインドルール（em_rule_xml/em_mapping_rule）を追加。w19ad系はCMDB登録も必要。",
    },
    "HW監視": {
        "priority": "方式調査",
        "cause": "nodeが全件trapSV（SNMP Trap集約サーバ）に集約されておりCMDB未登録。実機FQDNはadditional_info.alertSystemFQDNに存在。",
        "action": "u_transformation_rule で alertSystemFQDN を node に展開するルール追加。",
    },
    "syslog": {
        "priority": "方式調査",
        "cause": "type/resource全件空。gw系nodeにFPCサフィックス（-fpc0/-fpc1）が付きCMDB名と不一致。IPアドレス形式ノードはCMDB未登録。",
        "action": "node正規化（-fpc除去）。IPアドレス系CMDB登録。",
    },
    "CloudWatchLogs(HIOS)": {
        "priority": "高",
        "cause": "typeフィールドが全件空。CIバインドルールが未設定。BO-CAP_v6のみCMDB未登録。",
        "action": "CIバインドルール追加・BO-CAP_v6 CMDB登録。",
    },
    "Triplエラー": {
        "priority": "S-in後対応",
        "cause": "node/type/resource全件空。PandoraFMSのエラーレポートで additional_info.full_text にgw/sssw系NW機器名あり。",
        "action": "u_transformation_rule で full_text からnode展開。",
    },
    "RDS": {
        "priority": "中",
        "cause": "type/resource全件空。nodeはbig/bsd系DBインスタンス名（CMDB登録済み）。CIバインドルールが未設定。",
        "action": "RDSソース向けCIバインドルール設定。",
    },
    "Mackerel": {
        "priority": "対応不要",
        "cause": "node/type/resource全件空。additional_info.full_textにホスト名あり。全ホストCMDB未登録。",
        "action": "node展開ルール追加・CMDB登録（SLBGlobal/V10Global等）。",
    },
    "iMark_SV": {
        "priority": "中",
        "cause": "node=集約サーバ（trapSV/servicekansi）。additional_info から実機ノード名展開が必要。",
        "action": "additional_info から実機ノード名展開ルール追加。",
    },
}

_SHEET_NAME_MAX = 31


def _gs(rec: dict, fname: str) -> str:
    val = rec.get(fname, "")
    if isinstance(val, dict):
        return str(val.get("display_value") or val.get("value") or "")
    return str(val or "")


def _sheet_name(src: str) -> str:
    invalid = r'\/:*?"<>|'
    name = "".join(c for c in src if c not in invalid) or "unnamed"
    return name[:_SHEET_NAME_MAX]


def is_excluded(rec: dict) -> tuple[bool, str]:
    source = _gs(rec, "source")
    node   = _gs(rec, "node")
    if source in EXCLUDE_EXACT_SOURCES:
        return True, source
    if source.startswith("キャリア障害"):
        return True, "キャリア障害系"
    if source == "iMark_AWS" and "Servicekanshi" in node:
        return True, "iMark_AWS(Servicekanshi)"
    if source == "Zabbix" and "test-servicenow" in node:
        return True, "Zabbix正常性確認(ハートビート)"
    return False, ""


def fetch_unbound_alerts(token: str, max_records: int, extra_query: str = "") -> list[dict]:
    all_recs: list[dict] = []
    offset = 0
    base_query = "cmdb_ciISEMPTY"
    if extra_query:
        base_query = f"{base_query}^{extra_query}"
    while True:
        limit = PAGE_SIZE
        if max_records > 0:
            remaining = max_records - len(all_recs)
            if remaining <= 0:
                break
            limit = min(PAGE_SIZE, remaining)
        recs = snow_client.table_get(token, "em_alert", {
            "sysparm_limit":         limit,
            "sysparm_offset":        offset,
            "sysparm_query":         base_query,
            "sysparm_display_value": "true",
            "sysparm_fields":        ",".join(FETCH_FIELDS),
        })
        if not recs:
            break
        all_recs.extend(recs)
        print(f"  取得済み: {len(all_recs)} 件", file=sys.stderr)
        if len(recs) < PAGE_SIZE:
            break
        offset += PAGE_SIZE
    return all_recs


def lookup_cmdb_nodes(token: str, nodes: set[str]) -> dict[str, dict]:
    def _is_ip(s: str) -> bool:
        parts = s.split(".")
        return len(parts) == 4 and all(p.isdigit() for p in parts)

    valid = [n for n in nodes if n and not _is_ip(n)]
    if not valid:
        return {}

    result: dict[str, dict] = {}
    chunk_size = 80
    print(f"  CMDB照合: {len(valid)} ノード → ", file=sys.stderr, end="", flush=True)
    for i in range(0, len(valid), chunk_size):
        chunk = valid[i:i + chunk_size]
        escaped = [n.replace(",", r"\,") for n in chunk]
        try:
            recs = snow_client.table_get(token, "cmdb_ci", {
                "sysparm_limit":         500,
                "sysparm_query":         "nameIN" + ",".join(escaped),
                "sysparm_display_value": "true",
                "sysparm_fields":        "sys_id,name,sys_class_name,ip_address",
            })
            for rec in recs:
                name = _gs(rec, "name")
                if name:
                    result[name] = rec
        except Exception as exc:
            print(f"\n  CMDB照合エラー: {exc}", file=sys.stderr)
    print(f"{len(result)} 件ヒット", file=sys.stderr)
    return result


# ---------------------------------------------------------------------------
# Excel 出力
# ---------------------------------------------------------------------------

def output_excel(
    generated_at: str,
    all_recs: list[dict],
    required_recs: list[dict],
    by_source: dict[str, list[dict]],
    excluded_by_reason: dict[str, int],
    cmdb_maps: dict[str, dict],   # source → {node: cmdb_rec}
    filepath: str,
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl が必要です: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()

    # サンプルファイルに合わせた色定義
    title_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")  # 紺
    hdr_fill   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 青
    info_fill  = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")  # 黄
    prio_fill  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # 橙
    ok_fill    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 緑
    ng_fill    = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")  # 赤

    title_font = Font(color="FFFFFF", bold=True, size=13)
    hdr_font   = Font(color="FFFFFF", bold=True)
    wrap       = Alignment(wrap_text=True, vertical="top")
    center     = Alignment(horizontal="center")

    def _hdr(ws, row: int, col: int, val: str) -> None:
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center

    def _info_cell(ws, row: int, col: int, val) -> None:
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = info_fill
        cell.alignment = wrap

    def _prio_cell(ws, row: int, col: int, val: str) -> None:
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = prio_fill

    def _color_exist(cell, val: str) -> None:
        if val == "あり":
            cell.fill = ok_fill
        elif val == "なし":
            cell.fill = ng_fill

    # ================================================================
    # 1. 概要シート
    # ================================================================
    ws0 = wb.active
    ws0.title = "概要"

    cell = ws0.cell(row=1, column=1, value="CI未バインド アラート 分析レポート")
    cell.fill = title_fill
    cell.font = Font(color="FFFFFF", bold=True, size=14)

    ws0.cell(row=2, column=1, value="取得日時")
    ws0.cell(row=2, column=2, value=generated_at)
    ws0.cell(row=3, column=1, value="CI未バインド総件数（除外前）")
    ws0.cell(row=3, column=2, value=len(all_recs))
    ws0.cell(row=4, column=1, value="除外件数")
    ws0.cell(row=4, column=2, value=len(all_recs) - len(required_recs))
    ws0.cell(row=5, column=1, value="要対応件数")
    ws0.cell(row=5, column=2, value=len(required_recs))
    ws0["A5"].font = Font(bold=True)

    ws0.cell(row=7, column=1, value="除外条件一覧").font = Font(bold=True)
    _hdr(ws0, 8, 1, "除外理由")
    _hdr(ws0, 8, 2, "件数")
    for r_idx, (reason, count) in enumerate(
            sorted(excluded_by_reason.items(), key=lambda x: -x[1]), 9):
        ws0.cell(row=r_idx, column=1, value=reason)
        ws0.cell(row=r_idx, column=2, value=count)

    excl_end = 9 + len(excluded_by_reason)
    ws0.cell(row=excl_end + 1, column=1,
             value="source別 要対応件数").font = Font(bold=True)
    ws0.cell(row=excl_end + 1, column=3, value="バインド設定を入れるかどうか")
    _hdr(ws0, excl_end + 2, 1, "source")
    _hdr(ws0, excl_end + 2, 2, "件数")
    _hdr(ws0, excl_end + 2, 3, "優先度")
    _hdr(ws0, excl_end + 2, 4, "原因（概要）")
    for r_idx, (src, recs) in enumerate(
            sorted(by_source.items(), key=lambda x: -len(x[1])),
            excl_end + 3):
        ann = SOURCE_ANNOTATIONS.get(src, {})
        ws0.cell(row=r_idx, column=1, value=src)
        ws0.cell(row=r_idx, column=2, value=len(recs))
        ws0.cell(row=r_idx, column=3, value=ann.get("priority", ""))
        c = ws0.cell(row=r_idx, column=4, value=ann.get("cause", ""))
        c.alignment = wrap

    ws0.column_dimensions["A"].width = 40
    ws0.column_dimensions["B"].width = 10
    ws0.column_dimensions["C"].width = 18
    ws0.column_dimensions["D"].width = 60

    # ================================================================
    # 2. CI未バインド一覧シート
    # ================================================================
    ws_list = wb.create_sheet("CI未バインド一覧")
    list_cols = [
        "source", "node", "type", "resource", "severity",
        "classification", "u_type_category", "u_matched_rules",
        "sys_created_on", "sys_id",
    ]
    for col, h in enumerate(list_cols, 1):
        _hdr(ws_list, 1, col, h)
    for r_idx, rec in enumerate(required_recs, 2):
        for col, fname in enumerate(list_cols, 1):
            val = _gs(rec, fname)
            ws_list.cell(row=r_idx, column=col, value=val[:500] if val else "")

    ws_list.column_dimensions["A"].width = 12
    ws_list.column_dimensions["B"].width = 41
    ws_list.column_dimensions["C"].width = 55
    ws_list.column_dimensions["D"].width = 16
    ws_list.column_dimensions["E"].width = 9
    ws_list.column_dimensions["F"].width = 14
    ws_list.column_dimensions["G"].width = 14
    ws_list.column_dimensions["H"].width = 53
    ws_list.column_dimensions["I"].width = 14
    ws_list.column_dimensions["J"].width = 41

    # ================================================================
    # 3. ソース別シート
    # ================================================================
    for src, recs in sorted(by_source.items(), key=lambda x: -len(x[1])):
        sheet_title = _sheet_name(src)
        existing = {s.title for s in wb.worksheets}
        if sheet_title in existing:
            sheet_title = sheet_title[:28] + "_2"
        ws = wb.create_sheet(sheet_title)

        ann      = SOURCE_ANNOTATIONS.get(src, {})
        cmdb_map = cmdb_maps.get(src, {})
        unique_nodes = len({_gs(r, "node") for r in recs if _gs(r, "node")})

        # ソースシート列幅
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 45
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 11
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 11
        ws.column_dimensions["H"].width = 8
        ws.column_dimensions["I"].width = 55
        ws.column_dimensions["J"].width = 8

        # ---- タイトル行（紺）----
        cell = ws.cell(row=1, column=1, value=f"【{src}】 CI未バインド 分析結果")
        cell.fill = title_fill
        cell.font = title_font

        # ---- 情報セクション（row 3〜10）----
        ws.cell(row=3, column=1, value="件数")
        ws.cell(row=3, column=2, value=len(recs))

        ws.cell(row=4, column=1, value="node種類")
        ws.cell(row=4, column=2, value=f"{unique_nodes}種")

        ws.cell(row=5, column=1, value="CMDB照合（上位20）")
        if src in CMDB_LOOKUP_SOURCES and cmdb_map:
            found = len({n for n in {_gs(r, "node") for r in recs} if cmdb_map.get(n)})
            _info_cell(ws, 5, 2, f"登録済 {found}種 / 未登録 {unique_nodes - found}種")
        else:
            _info_cell(ws, 5, 2, "未実施")

        ws.cell(row=6, column=1, value="対応優先度")
        _prio_cell(ws, 6, 2, ann.get("priority", ""))

        ws.cell(row=8, column=1, value="未バインド原因")
        _info_cell(ws, 8, 2, ann.get("cause", ""))

        ws.cell(row=9, column=1, value="対応方針")
        _info_cell(ws, 9, 2, ann.get("action", ""))

        ws.cell(row=10, column=1, value="CMDB登録状況")
        if src in CMDB_LOOKUP_SOURCES and cmdb_map:
            found = len({n for n in {_gs(r, "node") for r in recs} if cmdb_map.get(n)})
            _info_cell(ws, 10, 2,
                       f"登録済 {found}種 / 未登録 {unique_nodes - found}種"
                       f"（詳細は下部のCMDB照合表を参照）")
        else:
            _info_cell(ws, 10, 2, "")

        # ---- additional_info サンプル（最大2件）----
        ws.cell(row=12, column=1, value="additional_info サンプル").font = Font(bold=True)
        sample_row = 13
        samples_written = 0
        for rec in recs:
            if samples_written >= 2:
                break
            ai = _gs(rec, "additional_info")
            if not ai:
                continue
            node  = _gs(rec, "node")
            type_ = _gs(rec, "type")
            sev   = _gs(rec, "severity")
            ws.cell(row=sample_row, column=1,
                    value=f"node='{node}'  type='{type_}'  sev='{sev}'")
            ws.cell(row=sample_row + 1, column=1, value=ai[:500])
            sample_row += 2
            samples_written += 1

        # ---- 分布表 ----
        dist_row = sample_row + 1

        node_cnt     = Counter(_gs(r, "node")     or "(空)" for r in recs)
        type_cnt     = Counter(_gs(r, "type")     or "(空)" for r in recs)
        resource_cnt = Counter(_gs(r, "resource") or "(空)" for r in recs)
        severity_cnt = Counter(_gs(r, "severity") or "(空)" for r in recs)
        mr_cnt: Counter = Counter()
        for r in recs:
            for rule in _gs(r, "u_matched_rules").split("\n"):
                rule = rule.strip()
                if rule:
                    mr_cnt[rule] += 1

        node_items     = node_cnt.most_common()
        type_items     = type_cnt.most_common()
        resource_items = resource_cnt.most_common()
        severity_items = severity_cnt.most_common()
        mr_items       = mr_cnt.most_common(50)
        max_dist = max(len(node_items), len(type_items),
                       len(resource_items), len(severity_items), len(mr_items), 1)

        _hdr(ws, dist_row, 1, f"node別（{len(node_cnt)}種）")
        _hdr(ws, dist_row, 2, "件数")
        _hdr(ws, dist_row, 3, f"type別（{len(type_cnt)}種）")
        _hdr(ws, dist_row, 4, "件数")
        _hdr(ws, dist_row, 5, "resource別")
        _hdr(ws, dist_row, 6, "件数")
        _hdr(ws, dist_row, 7, "severity別")
        _hdr(ws, dist_row, 8, "件数")
        _hdr(ws, dist_row, 9, "u_matched_rules")
        _hdr(ws, dist_row, 10, "件数")

        for i in range(max_dist):
            row = dist_row + 1 + i
            if i < len(node_items):
                ws.cell(row=row, column=1, value=node_items[i][0])
                ws.cell(row=row, column=2, value=node_items[i][1])
            if i < len(type_items):
                ws.cell(row=row, column=3, value=type_items[i][0])
                ws.cell(row=row, column=4, value=type_items[i][1])
            if i < len(resource_items):
                ws.cell(row=row, column=5, value=resource_items[i][0])
                ws.cell(row=row, column=6, value=resource_items[i][1])
            if i < len(severity_items):
                ws.cell(row=row, column=7, value=severity_items[i][0])
                ws.cell(row=row, column=8, value=severity_items[i][1])
            if i < len(mr_items):
                ws.cell(row=row, column=9, value=mr_items[i][0][:150])
                ws.cell(row=row, column=10, value=mr_items[i][1])

        # ---- 疑義CI CMDB照合表（3ソースのみ）----
        if src in CMDB_LOOKUP_SOURCES:
            cmdb_tbl_row = dist_row + 1 + max_dist + 3

            cell = ws.cell(row=cmdb_tbl_row, column=1,
                           value=f"【疑義CI CMDB照合表】（調査日: {generated_at[:10]}）")
            cell.fill = title_fill
            cell.font = title_font

            _hdr(ws, cmdb_tbl_row + 1, 1, "ノード名")
            _hdr(ws, cmdb_tbl_row + 1, 2, "アラート件数")
            _hdr(ws, cmdb_tbl_row + 1, 3, "CMDB_存在")
            _hdr(ws, cmdb_tbl_row + 1, 4, "CMDBクラス")
            _hdr(ws, cmdb_tbl_row + 1, 5, "CMDB_sys_id")

            node_summary: dict[str, dict] = {}
            for rec in recs:
                node = _gs(rec, "node") or "(空)"
                if node not in node_summary:
                    ci = cmdb_map.get(node) if node != "(空)" else None
                    node_summary[node] = {
                        "count":       0,
                        "cmdb_exist":  "あり" if ci else ("なし" if node != "(空)" else ""),
                        "cmdb_class":  _gs(ci, "sys_class_name") if ci else "",
                        "cmdb_sys_id": _gs(ci, "sys_id") if ci else "",
                    }
                node_summary[node]["count"] += 1

            for r_idx, (node, stat) in enumerate(
                    sorted(node_summary.items(), key=lambda x: -x[1]["count"]),
                    cmdb_tbl_row + 2):
                ws.cell(row=r_idx, column=1, value=node)
                ws.cell(row=r_idx, column=2, value=stat["count"])
                cell_exist = ws.cell(row=r_idx, column=3, value=stat["cmdb_exist"])
                _color_exist(cell_exist, stat["cmdb_exist"])
                ws.cell(row=r_idx, column=4, value=stat["cmdb_class"])
                ws.cell(row=r_idx, column=5, value=stat["cmdb_sys_id"])

    # ================================================================
    # N. 除外ソースシート
    # ================================================================
    ws_ex = wb.create_sheet("除外ソース")
    _hdr(ws_ex, 1, 1, "除外理由")
    _hdr(ws_ex, 1, 2, "件数")
    for r_idx, (reason, count) in enumerate(
            sorted(excluded_by_reason.items(), key=lambda x: -x[1]), 2):
        ws_ex.cell(row=r_idx, column=1, value=reason)
        ws_ex.cell(row=r_idx, column=2, value=count)
    ws_ex.column_dimensions["A"].width = 40
    ws_ex.column_dimensions["B"].width = 10

    wb.save(filepath)
    print(f"Excel 出力完了: {filepath}", file=sys.stderr)


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="CI未バインドアラート調査（CMDB照合付き）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--file", default=None,
                        help="出力 Excel ファイルパス (default: tmpdir/ci_unbound_alerts_YYYYMMDD_HHMM.xlsx)")
    parser.add_argument("--limit", type=int, default=0,
                        help="取得上限件数（0=全件, default: 0）")
    parser.add_argument("--output", choices=["excel", "json"], default="excel",
                        help="出力形式 (default: excel)")
    parser.add_argument("--no-cmdb", action="store_true",
                        help="CMDB照合をスキップする（高速確認用）")
    parser.add_argument("--query", default="",
                        help="追加 sysparm_query フィルタ（例: sys_created_on>=2026-04-01^sys_created_on<2026-06-19）")
    args = parser.parse_args()

    import datetime
    now = datetime.datetime.now()
    generated_at = now.strftime("%Y-%m-%d %H:%M:%S")
    if args.file is None:
        args.file = now.strftime("tmpdir/ci_unbound_alerts_%Y%m%d.xlsx")

    print("OAuth トークンを取得中...", file=sys.stderr)
    try:
        token = snow_client.get_token()
    except Exception as exc:
        print(f"認証失敗: {exc}", file=sys.stderr)
        sys.exit(1)

    print("CI未バインドアラートを取得中...", file=sys.stderr)
    all_recs = fetch_unbound_alerts(token, args.limit, args.query)
    print(f"  合計取得: {len(all_recs)} 件", file=sys.stderr)

    excluded_by_reason: dict[str, int] = defaultdict(int)
    required_recs: list[dict] = []
    for rec in all_recs:
        excl, reason = is_excluded(rec)
        if excl:
            excluded_by_reason[reason] += 1
        else:
            required_recs.append(rec)

    print(f"  除外: {sum(excluded_by_reason.values())} 件 / 要対応: {len(required_recs)} 件",
          file=sys.stderr)

    by_source: dict[str, list[dict]] = defaultdict(list)
    for rec in required_recs:
        by_source[_gs(rec, "source") or "(空)"].append(rec)

    # CMDB照合（Zabbix / CloudWatchLogs / HIOS(SV)）
    cmdb_maps: dict[str, dict] = {}
    if not args.no_cmdb:
        for src in CMDB_LOOKUP_SOURCES:
            src_recs = by_source.get(src, [])
            if not src_recs:
                print(f"  {src}: 0 件（スキップ）", file=sys.stderr)
                cmdb_maps[src] = {}
                continue
            print(f"  {src}: {len(src_recs)} 件 CMDB照合中...", file=sys.stderr)
            nodes = {_gs(r, "node") for r in src_recs if _gs(r, "node")}
            cmdb_maps[src] = lookup_cmdb_nodes(token, nodes)

    if args.output == "json":
        print(json.dumps({
            "generated_at":       generated_at,
            "total_count":        len(all_recs),
            "excluded_count":     sum(excluded_by_reason.values()),
            "required_count":     len(required_recs),
            "excluded_by_reason": dict(excluded_by_reason),
            "sources": {src: len(recs) for src, recs in by_source.items()},
        }, ensure_ascii=False, indent=2))
    else:
        output_excel(
            generated_at, all_recs, required_recs,
            by_source, dict(excluded_by_reason), cmdb_maps, args.file,
        )


if __name__ == "__main__":
    main()
