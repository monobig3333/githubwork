#!/usr/bin/env python3
"""
ServiceNow CI バインディングルール仕様書を生成する。

このインスタンスの CIバインディングルールは主に sa_event_rule に格納されている。
さらに以下のテーブルも対象:
  - sa_event_rule       : イベントルール（CIバインディング・フィールドマッピング等）← 主要
  - em_rule_xml         : イベントルール XML（matchRule = CIバインディング, fieldMappingRule = フィールドマッピング）
  - em_mapping_rule     : フィールドマッピングルール
  - em_binding_device_map: CI タイプ間のマッピング定義

使用例:
  python get_ci_bindings.py                      # JSON 出力（標準出力）
  python get_ci_bindings.py --output excel       # Excel ファイル出力
  python get_ci_bindings.py --output excel --file tmpdir/ci_binding_spec.xlsx
"""

import sys
import os
import json
import argparse
import xml.etree.ElementTree as ET

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# cichk は本番環境 (biglobeprod) を対象とする
os.environ["SNOW_BASE_URL"] = "https://biglobeprod.service-now.com"
os.environ["SNOW_SECRET_NAME"] = "servicenow/api/credentials/biglobeprod/admin-ai-api"
os.environ["SNOW_CLIENT_ID"] = ""
os.environ["SNOW_CLIENT_SECRET"] = ""

import snow_client


def _gs(rec: dict, fname: str) -> str:
    val = rec.get(fname, "")
    if isinstance(val, dict):
        return str(val.get("display_value") or val.get("value") or "")
    return str(val or "")


def _xt(root, path: str, default: str = "") -> str:
    el = root.find(path)
    return el.text.strip() if el is not None and el.text else default


def fetch_table(token: str, table: str, fields: list[str],
                query: str = "", limit: int = 500) -> list[dict]:
    try:
        return snow_client.table_get(token, table, {
            "sysparm_limit":         limit,
            "sysparm_query":         query,
            "sysparm_display_value": "true",
            "sysparm_fields":        ",".join(fields),
        })
    except Exception as exc:
        print(f"  [{table}] 取得失敗: {exc}", file=sys.stderr)
        return []


# ---------------------------------------------------------------------------
# em_rule_xml パーサー
# ---------------------------------------------------------------------------

def parse_rule_xml(rec: dict) -> dict | None:
    """em_rule_xml の1レコードを解析して構造化データを返す。"""
    sys_id = _gs(rec, "sys_id")
    fname  = _gs(rec, "name")
    xml_str = _gs(rec, "rule_xml")
    if not xml_str:
        return None
    try:
        root = ET.fromstring(xml_str)
    except ET.ParseError as e:
        print(f"  XML解析失敗 {fname}: {e}", file=sys.stderr)
        return None

    tag = root.tag

    if tag == "matchRule":
        # CIバインディングルール
        match_fields = []
        for mf in root.findall("matchFields"):
            field = _xt(mf, "field")
            regex = _xt(mf, "regex")
            groups = [g.text for g in mf.findall("groupMappings") if g.text]
            match_fields.append({"field": field, "regex": regex, "groups": groups})

        field_comps = []
        for fc in root.findall("fieldCompositions"):
            field_comps.append({
                "field": _xt(fc, "fieldName"),
                "template": _xt(fc, "composingFields"),
            })

        return {
            "rule_type":      "matchRule",
            "sys_id":         sys_id,
            "file_name":      fname,
            "name":           _xt(root, "displayName") or _xt(root, "name"),
            "source":         _xt(root, "ems"),
            "ci_type":        _xt(root, "ciTypeName"),
            "ci_type_id":     _xt(root, "ciTypeId"),
            "rule_order":     _xt(root, "ruleOrder"),
            "hidden":         _xt(root, "hidden"),
            "ignore":         _xt(root, "ignore"),
            "expire_minutes": _xt(root, "expireMinutes"),
            "match_fields":   match_fields,
            "field_compositions": field_comps,
        }

    elif tag == "fieldMappingRule":
        # フィールドマッピングルール
        const_mappings = []
        cm = root.find("constantMappings")
        if cm is not None:
            for entry in cm.findall("entry"):
                key = _xt(entry, "key")
                val = _xt(entry, "value")
                if key:
                    const_mappings.append({"key": key, "value": val})

        return {
            "rule_type":       "fieldMappingRule",
            "sys_id":          sys_id,
            "file_name":       fname,
            "name":            fname,
            "source":          _xt(root, "ems"),
            "constant_mappings": const_mappings,
        }

    else:
        return {
            "rule_type": tag,
            "sys_id":    sys_id,
            "file_name": fname,
            "name":      fname,
        }


# ---------------------------------------------------------------------------
# データ取得
# ---------------------------------------------------------------------------

SA_EVENT_RULE_FIELDS = [
    "sys_id", "name", "active", "order", "type", "description",
    "ems_source", "conditions", "ci_class", "script",
    "sys_created_on", "sys_updated_on",
]


def fetch_all(token: str) -> dict:
    print("sa_event_rule を取得中...", file=sys.stderr)
    sa_event_rules = fetch_table(token, "sa_event_rule", SA_EVENT_RULE_FIELDS,
                                 limit=2000)
    print(f"  sa_event_rule: {len(sa_event_rules)} 件", file=sys.stderr)

    print("em_rule_xml を取得・解析中...", file=sys.stderr)
    rule_xml_recs = fetch_table(token, "em_rule_xml",
                                ["sys_id", "name", "rule_xml"])
    parsed_rules = [r for rec in rule_xml_recs
                    if (r := parse_rule_xml(rec)) is not None]
    match_rules  = [r for r in parsed_rules if r["rule_type"] == "matchRule"]
    field_rules  = [r for r in parsed_rules if r["rule_type"] == "fieldMappingRule"]
    print(f"  matchRule (CIバインディング): {len(match_rules)} 件", file=sys.stderr)
    print(f"  fieldMappingRule: {len(field_rules)} 件", file=sys.stderr)

    print("em_mapping_rule を取得中...", file=sys.stderr)
    mapping_rules = fetch_table(token, "em_mapping_rule", [
        "sys_id", "name", "active", "order", "mapping_type",
        "event_rule", "from_field", "to_field", "alert_field",
        "filter", "value", "run_after_binding",
    ])
    print(f"  em_mapping_rule: {len(mapping_rules)} 件", file=sys.stderr)

    print("em_binding_device_map を取得中...", file=sys.stderr)
    device_map = fetch_table(token, "em_binding_device_map",
                             ["sys_id", "from_ci_type", "to_ci_type", "mapped_column"])
    print(f"  em_binding_device_map: {len(device_map)} 件", file=sys.stderr)

    print("u_transformation_rule を取得中...", file=sys.stderr)
    trans_rules = fetch_table(token, "u_transformation_rule", [
        "sys_id", "u_rule_name", "u_order", "u_category", "u_active",
        "u_table_name", "u_conditions", "u_description", "u_data_type",
    ], limit=2000)
    print(f"  u_transformation_rule: {len(trans_rules)} 件", file=sys.stderr)

    print("u_transformation_rule_detail を取得中...", file=sys.stderr)
    trans_details = fetch_table(token, "u_transformation_rule_detail", [
        "sys_id", "u_parent", "u_field_name", "u_table_name",
        "u_target_string", "u_transformed_string",
    ], limit=5000)
    print(f"  u_transformation_rule_detail: {len(trans_details)} 件", file=sys.stderr)

    print("u_ignore_rule を取得中...", file=sys.stderr)
    ignore_rules = fetch_table(token, "u_ignore_rule", [
        "sys_id", "u_number", "u_alert_type", "u_service_id", "u_device_name",
        "u_monitoring_number", "u_monitoring_type",
        "u_ignore_start_date", "u_ignore_end_date",
        "u_applicant", "u_department", "u_message", "u_change_request",
    ], limit=2000)
    print(f"  u_ignore_rule: {len(ignore_rules)} 件", file=sys.stderr)

    print("u_alert_type を取得中...", file=sys.stderr)
    alert_types = fetch_table(token, "u_alert_type",
                              ["sys_id", "u_name", "u_active"])
    print(f"  u_alert_type: {len(alert_types)} 件", file=sys.stderr)

    return {
        "sa_event_rules":  sa_event_rules,
        "match_rules":     match_rules,
        "field_rules_xml": field_rules,
        "mapping_rules":   mapping_rules,
        "device_map":      device_map,
        "trans_rules":     trans_rules,
        "trans_details":   trans_details,
        "ignore_rules":    ignore_rules,
        "alert_types":     alert_types,
    }


# ---------------------------------------------------------------------------
# 出力
# ---------------------------------------------------------------------------

def output_json(data: dict) -> None:
    print(json.dumps(data, ensure_ascii=False, indent=2))


def output_excel(data: dict, filepath: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl が必要です: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def write_header(ws, headers: list[str]) -> None:
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    # ---- 概要シート ----
    ws0 = wb.active
    ws0.title = "概要"
    ws0["A1"] = "CI バインディングルール 仕様書"
    ws0["A1"].font = Font(size=14, bold=True)
    ws0["A3"] = "取得元テーブル:"
    ws0["A4"]  = "sa_event_rule";         ws0["B4"]  = "イベントルール（CIバインディング等）← 主要"
    ws0["A5"]  = "em_rule_xml";           ws0["B5"]  = "イベントルール XML（matchRule/fieldMappingRule）"
    ws0["A6"]  = "em_mapping_rule";       ws0["B6"]  = "フィールドマッピングルール"
    ws0["A7"]  = "em_binding_device_map"; ws0["B7"]  = "CI タイプ間マッピング"
    ws0["A8"]  = "u_transformation_rule"; ws0["B8"]  = "カスタムアラート変換ルール（主要）"
    ws0["A9"]  = "u_transformation_rule_detail"; ws0["B9"] = "アラート変換ルール 詳細"
    ws0["A10"] = "u_ignore_rule";         ws0["B10"] = "アラート無視ルール"
    ws0["A11"] = "u_alert_type";          ws0["B11"] = "アラートタイプ定義"
    ws0["A13"] = "件数:"
    ws0["A14"] = "sa_event_rule 件数:";                         ws0["B14"] = len(data["sa_event_rules"])
    ws0["A15"] = "matchRule（CIバインディング）件数:";           ws0["B15"] = len(data["match_rules"])
    ws0["A16"] = "fieldMappingRule（XML）件数:";                ws0["B16"] = len(data["field_rules_xml"])
    ws0["A17"] = "em_mapping_rule 件数:";                       ws0["B17"] = len(data["mapping_rules"])
    ws0["A18"] = "em_binding_device_map 件数:";                 ws0["B18"] = len(data["device_map"])
    ws0["A19"] = "u_transformation_rule 件数:";                 ws0["B19"] = len(data["trans_rules"])
    ws0["A20"] = "u_transformation_rule_detail 件数:";          ws0["B20"] = len(data["trans_details"])
    ws0["A21"] = "u_ignore_rule 件数:";                         ws0["B21"] = len(data["ignore_rules"])
    ws0["A22"] = "u_alert_type 件数:";                          ws0["B22"] = len(data["alert_types"])
    ws0["A24"] = "※ このファイルは ServiceNow 本番環境 (biglobeprod) の実データから自動生成しました。"
    ws0["A25"] = "※ 大半の CIバインドルールは sa_event_rule に格納されています。"

    # ---- sa_event_rule シート ----
    ws_sa = wb.create_sheet("イベントルール(sa_event_rule)")
    headers_sa = [
        "ルール名", "有効", "順序", "タイプ", "イベントソース", "CIクラス",
        "説明", "条件", "スクリプト有無", "作成日時", "更新日時", "sys_id",
    ]
    write_header(ws_sa, headers_sa)
    for row_idx, rec in enumerate(data["sa_event_rules"], 2):
        script = _gs(rec, "script")
        ws_sa.cell(row=row_idx, column=1,  value=_gs(rec, "name"))
        ws_sa.cell(row=row_idx, column=2,  value=_gs(rec, "active"))
        ws_sa.cell(row=row_idx, column=3,  value=_gs(rec, "order"))
        ws_sa.cell(row=row_idx, column=4,  value=_gs(rec, "type"))
        ws_sa.cell(row=row_idx, column=5,  value=_gs(rec, "ems_source"))
        ws_sa.cell(row=row_idx, column=6,  value=_gs(rec, "ci_class"))
        ws_sa.cell(row=row_idx, column=7,  value=_gs(rec, "description")[:200])
        ws_sa.cell(row=row_idx, column=8,  value=_gs(rec, "conditions")[:200])
        ws_sa.cell(row=row_idx, column=9,  value="あり" if script else "")
        ws_sa.cell(row=row_idx, column=10, value=_gs(rec, "sys_created_on"))
        ws_sa.cell(row=row_idx, column=11, value=_gs(rec, "sys_updated_on"))
        ws_sa.cell(row=row_idx, column=12, value=_gs(rec, "sys_id"))

    # ---- CIバインディングルール（matchRule）シート ----
    ws1 = wb.create_sheet("CIバインディングルール")
    headers1 = [
        "ルール名", "ソース(EMS)", "CIタイプ", "順序", "hidden", "ignore",
        "有効期限(分)", "マッチフィールド数", "フィールド組み立て数", "sys_id",
    ]
    write_header(ws1, headers1)
    for row_idx, r in enumerate(data["match_rules"], 2):
        ws1.cell(row=row_idx, column=1,  value=r.get("name", ""))
        ws1.cell(row=row_idx, column=2,  value=r.get("source", ""))
        ws1.cell(row=row_idx, column=3,  value=r.get("ci_type", ""))
        ws1.cell(row=row_idx, column=4,  value=r.get("rule_order", ""))
        ws1.cell(row=row_idx, column=5,  value=r.get("hidden", ""))
        ws1.cell(row=row_idx, column=6,  value=r.get("ignore", ""))
        ws1.cell(row=row_idx, column=7,  value=r.get("expire_minutes", ""))
        ws1.cell(row=row_idx, column=8,  value=len(r.get("match_fields", [])))
        ws1.cell(row=row_idx, column=9,  value=len(r.get("field_compositions", [])))
        ws1.cell(row=row_idx, column=10, value=r.get("sys_id", ""))

    # ---- CIバインディング マッチフィールド詳細シート ----
    ws1d = wb.create_sheet("CIバインディング_マッチフィールド")
    headers1d = ["ルール名", "ソース", "CIタイプ", "フィールド", "正規表現", "グループマッピング", "sys_id"]
    write_header(ws1d, headers1d)
    row_idx = 2
    for r in data["match_rules"]:
        for mf in r.get("match_fields", []):
            ws1d.cell(row=row_idx, column=1, value=r.get("name", ""))
            ws1d.cell(row=row_idx, column=2, value=r.get("source", ""))
            ws1d.cell(row=row_idx, column=3, value=r.get("ci_type", ""))
            ws1d.cell(row=row_idx, column=4, value=mf.get("field", ""))
            ws1d.cell(row=row_idx, column=5, value=mf.get("regex", ""))
            ws1d.cell(row=row_idx, column=6, value=", ".join(mf.get("groups", [])))
            ws1d.cell(row=row_idx, column=7, value=r.get("sys_id", ""))
            row_idx += 1

    # ---- フィールドマッピングルール（em_mapping_rule）シート ----
    ws2 = wb.create_sheet("フィールドマッピングルール")
    headers2 = [
        "名前", "アクティブ", "順序", "マッピングタイプ",
        "元フィールド", "先フィールド", "アラートフィールド",
        "イベントルール参照", "バインディング後実行", "フィルター", "sys_id",
    ]
    write_header(ws2, headers2)
    for row_idx, rec in enumerate(data["mapping_rules"], 2):
        ws2.cell(row=row_idx, column=1,  value=_gs(rec, "name"))
        ws2.cell(row=row_idx, column=2,  value=_gs(rec, "active"))
        ws2.cell(row=row_idx, column=3,  value=_gs(rec, "order"))
        ws2.cell(row=row_idx, column=4,  value=_gs(rec, "mapping_type"))
        ws2.cell(row=row_idx, column=5,  value=_gs(rec, "from_field"))
        ws2.cell(row=row_idx, column=6,  value=_gs(rec, "to_field"))
        ws2.cell(row=row_idx, column=7,  value=_gs(rec, "alert_field"))
        ws2.cell(row=row_idx, column=8,  value=_gs(rec, "event_rule"))
        ws2.cell(row=row_idx, column=9,  value=_gs(rec, "run_after_binding"))
        ws2.cell(row=row_idx, column=10, value=_gs(rec, "filter")[:200])
        ws2.cell(row=row_idx, column=11, value=_gs(rec, "sys_id"))

    # ---- CIタイプマッピング（em_binding_device_map）シート ----
    ws3 = wb.create_sheet("CIタイプマッピング")
    headers3 = ["変換元CIタイプ", "変換先CIタイプ", "マッピングカラム", "sys_id"]
    write_header(ws3, headers3)
    for row_idx, rec in enumerate(data["device_map"], 2):
        ws3.cell(row=row_idx, column=1, value=_gs(rec, "from_ci_type"))
        ws3.cell(row=row_idx, column=2, value=_gs(rec, "to_ci_type"))
        ws3.cell(row=row_idx, column=3, value=_gs(rec, "mapped_column"))
        ws3.cell(row=row_idx, column=4, value=_gs(rec, "sys_id"))

    # ---- アラート変換ルール（u_transformation_rule）シート ----
    ws4 = wb.create_sheet("アラート変換ルール")
    headers4 = [
        "ルール名", "カテゴリ", "順序", "有効", "対象テーブル",
        "説明", "データ型", "条件", "sys_id",
    ]
    write_header(ws4, headers4)
    for row_idx, rec in enumerate(data["trans_rules"], 2):
        ws4.cell(row=row_idx, column=1, value=_gs(rec, "u_rule_name"))
        ws4.cell(row=row_idx, column=2, value=_gs(rec, "u_category"))
        ws4.cell(row=row_idx, column=3, value=_gs(rec, "u_order"))
        ws4.cell(row=row_idx, column=4, value=_gs(rec, "u_active"))
        ws4.cell(row=row_idx, column=5, value=_gs(rec, "u_table_name"))
        ws4.cell(row=row_idx, column=6, value=_gs(rec, "u_description"))
        ws4.cell(row=row_idx, column=7, value=_gs(rec, "u_data_type"))
        ws4.cell(row=row_idx, column=8, value=_gs(rec, "u_conditions")[:300])
        ws4.cell(row=row_idx, column=9, value=_gs(rec, "sys_id"))

    # ---- アラート変換ルール詳細（u_transformation_rule_detail）シート ----
    ws5 = wb.create_sheet("アラート変換ルール詳細")
    headers5 = [
        "親ルール名", "フィールド名", "対象テーブル",
        "変換元文字列", "変換後文字列", "sys_id",
    ]
    write_header(ws5, headers5)
    for row_idx, rec in enumerate(data["trans_details"], 2):
        ws5.cell(row=row_idx, column=1, value=_gs(rec, "u_parent"))
        ws5.cell(row=row_idx, column=2, value=_gs(rec, "u_field_name"))
        ws5.cell(row=row_idx, column=3, value=_gs(rec, "u_table_name"))
        ws5.cell(row=row_idx, column=4, value=_gs(rec, "u_target_string")[:200])
        ws5.cell(row=row_idx, column=5, value=_gs(rec, "u_transformed_string")[:200])
        ws5.cell(row=row_idx, column=6, value=_gs(rec, "sys_id"))

    # ---- アラート無視ルール（u_ignore_rule）シート ----
    ws6 = wb.create_sheet("アラート無視ルール")
    headers6 = [
        "番号", "アラートタイプ", "サービスID", "デバイス名",
        "監視番号", "監視種別", "無視開始日時", "無視終了日時",
        "申請者", "部門", "メッセージ", "変更リクエスト", "sys_id",
    ]
    write_header(ws6, headers6)
    for row_idx, rec in enumerate(data["ignore_rules"], 2):
        ws6.cell(row=row_idx, column=1,  value=_gs(rec, "u_number"))
        ws6.cell(row=row_idx, column=2,  value=_gs(rec, "u_alert_type"))
        ws6.cell(row=row_idx, column=3,  value=_gs(rec, "u_service_id"))
        ws6.cell(row=row_idx, column=4,  value=_gs(rec, "u_device_name")[:200])
        ws6.cell(row=row_idx, column=5,  value=_gs(rec, "u_monitoring_number"))
        ws6.cell(row=row_idx, column=6,  value=_gs(rec, "u_monitoring_type"))
        ws6.cell(row=row_idx, column=7,  value=_gs(rec, "u_ignore_start_date"))
        ws6.cell(row=row_idx, column=8,  value=_gs(rec, "u_ignore_end_date"))
        ws6.cell(row=row_idx, column=9,  value=_gs(rec, "u_applicant"))
        ws6.cell(row=row_idx, column=10, value=_gs(rec, "u_department"))
        ws6.cell(row=row_idx, column=11, value=_gs(rec, "u_message")[:200])
        ws6.cell(row=row_idx, column=12, value=_gs(rec, "u_change_request"))
        ws6.cell(row=row_idx, column=13, value=_gs(rec, "sys_id"))

    # ---- アラートタイプ（u_alert_type）シート ----
    ws7 = wb.create_sheet("アラートタイプ")
    headers7 = ["アラートタイプ名", "有効", "sys_id"]
    write_header(ws7, headers7)
    for row_idx, rec in enumerate(data["alert_types"], 2):
        ws7.cell(row=row_idx, column=1, value=_gs(rec, "u_name"))
        ws7.cell(row=row_idx, column=2, value=_gs(rec, "u_active"))
        ws7.cell(row=row_idx, column=3, value=_gs(rec, "sys_id"))

    wb.save(filepath)
    print(f"Excel 出力完了: {filepath}", file=sys.stderr)


# ---------------------------------------------------------------------------
# エントリポイント
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="ServiceNow CI バインディングルール仕様書を生成する",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--output", choices=["json", "excel"], default="json",
                        help="出力形式 json/excel (default: json)")
    parser.add_argument("--file", default="ci_binding_spec.xlsx",
                        help="Excel 出力ファイル名 (default: ci_binding_spec.xlsx)")
    args = parser.parse_args()

    print("OAuth トークンを取得中...", file=sys.stderr)
    try:
        token = snow_client.get_token()
    except Exception as exc:
        print(f"認証失敗: {exc}", file=sys.stderr)
        sys.exit(1)

    data = fetch_all(token)

    if args.output == "excel":
        output_excel(data, args.file)
    else:
        output_json(data)


if __name__ == "__main__":
    main()
